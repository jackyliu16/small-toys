"""
@Target : 
@Annotation : 
@Author : JackyLiu
@github : https://github.com/jackyliu16
@Date   : 2022/11/3
@Reference:
    https://blog.csdn.net/weixin_47971206/article/details/119900031
"""
from openpyxl import load_workbook
from openpyxl import Workbook
from typing import List, Dict


def get_information_from_txt(file_name: str, control_list: Dict[str, str]) -> List[Dict[str, str]]:
    with open(file_name, 'r', encoding='utf-8') as File:
        All_Data = []
        lines = File.readlines()
        # print("all data: ", end="")
        # print(lines)
        flag = False    # flag of if it's in a block
        # 这个代表了一个人的字典
        people = {}
        # 这个是可能存在的列标题【需要将长的放在前面方便匹配，否则会导致额外的信息】
        # list = ['姓名', '单位', '微信id', '微信', 'github id', 'github', '个人目标（期望）', '个人目标']
        key_list = list(control_list.keys())
        for line in lines:
            if line == "Choose a reason\n":
                if flag:   # if now in block => go out this block
                    print(f"people: {people}")
                    All_Data.append(people)
                    people = {}
                else:          # it's the first block
                    flag = True  # go into a block
            elif line.find("Hide comment") != -1:
                pass
            # 这个地方可以添加空行判断
            else:
                # 如果存在分隔符的情况
                if line.find(":") != -1 or line.find("：") != -1:
                    # 这个地方实际上是因为很多人符号有点问题外加上他的那个东西会被split错误识别，最后直接整了切片
                    # split = line.split(":")
                    # if len(split) != 2:
                    #     print("="*20+"Error"+"="*20)
                    #     print(f"line: {line}", end="")
                    #     print(f"split: {split}")
                    #     print("="*55)
                    # else:
                    #     people[split[0]] = split[1].strip()

                    # 这坨起到的作用在于找到其中可能存在的那个分隔符
                    idx = line.find(":")
                    split = line.split(":")
                    if idx == -1:
                        idx = line.find("：")
                        split = line.split("：")

                    # print(f"idx: {idx}")
                    # print(f"split: {split}")

                    # 差错判断
                    if len(split) < 2:
                        print("Error:")
                        print(f"line{line}")

                    # 判断前面是不是加了什么东西
                    for i in range(0, len(key_list)):
                        if line.find(key_list[i]) != -1:
                            # 如果本行存在这个元素
                            # print(f"len: {line.split(item)}")
                            people[key_list[i]] = line[idx +
                                                   1:].split(key_list[i])[-1].strip()
                            break
                        # TODO 采用这套鉴别操作很会产生一大堆报错
                        # if i == len(list)-1 :
                        #     print(f"line: {line}")
                    # people[split[0].strip()] = line[idx+1:].strip()
                # 如果不存在分隔符
                else:
                    # 对于一些没写分隔符的人，就只能尝试通过关键字识别，然后直接对着关键词来切【这个地方可能还要考虑中间有没有分隔符的情况】
                    for i in range(0, len(key_list)):
                        if line.find(key_list[i]) != -1:
                            # 如果本行存在这个元素
                            # print(f"len: {line.split(item)}")
                            people[key_list[i]] = line.split(key_list[i])[-1].strip()
                            break
                        # TODO 采用这套鉴别操作很会产生一大堆报错
                        # if i == len(list)-1 :
                        #     print(f"line: {line}")

        if flag:  # if now in block => go out this block
            print(f"people: {people}")
            All_Data.append(people)
            people = {}

        # print(All_Data)
        return All_Data


def put_data_in_xlsx(list: List[Dict[str, str]], control_list: Dict[str, str]):
    excel_path = "data_process_1/output.xlsx"

    wb = Workbook(excel_path)
    wb.save(excel_path)

    wb = load_workbook(excel_path)
    wb.create_sheet("output")
    ws = wb.active

    tmp = []
    for (k, v) in control_list.items():
        if v not in tmp:
            tmp.append(v)

    # table name
    for i in range(0, len(tmp)):
        ws.cell(1, i+1).value = tmp[i]

    for i in range(0, len(list)):
        for (k, v) in list[i].items():
            # 这个主要是为了列号
            ws.cell(i+2, tmp.index(control_list.get(k))+1).value = v 
    
    wb.save(excel_path)
    wb.close()



if __name__ == "__main__":
    # list = ['姓名', '单位', '微信id', '微信', 'github id', 'github', '个人目标（期望）', '个人目标']

    control_list = {
        '姓名':     '姓名',
        '单位':     '单位',
        "微信":     '微信id',
        "微信id":   '微信id',
        '电邮':     '电邮',
        "github id":'github id',
        'github':   'github id',
        '个人目标（期望）': '个人目标',
        '个人目标': '个人目标',
    }
    All_Data = get_information_from_txt(
        "./data_process_1/issue2.txt", control_list)
    put_data_in_xlsx(All_Data, control_list)
    print('最后，请注意：输入的数据是以Choose a reason为分节符的部分, 这就意味着你要删除无用的其他部分和所有与control_list无关的部分')
