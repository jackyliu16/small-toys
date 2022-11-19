#!/usr/bin/python
# -*- encoding: utf-8 -*-
'''
@File    :   generate_sql_from_csv.py
@Time    :   2022/11/18 14:02:41
@Author  :   jackyliu
@Version :   1.0
@Contact :   18922251299@163.com
@github  :   https://github.com/jackyliu16
@reference: 
    https://blog.csdn.net/qq_44614026/article/details/108083958#:~:text=openpyxl%E7%AE%80%E4%BB%8B%201%20openpyxl%20%E6%9C%80%E5%A5%BD%E7%94%A8%E7%9A%84%20python%20%E6%93%8D%E4%BD%9C%20excel%20%E8%A1%A8%E6%A0%BC%E5%BA%93%EF%BC%8C%E4%BD%86%E4%B8%8D%E6%98%AF%E5%AE%98%E6%96%B9%E7%9A%84%E6%A0%87%E5%87%86%E5%BA%93%EF%BC%8C%E9%9C%80%E8%A6%81%E6%89%8B%E5%8A%A8%E5%AE%89%E8%A3%85,2%20%E5%8F%AF%E4%BB%A5%E8%AF%BB%E5%8F%96%E5%92%8C%E5%86%99%E5%85%A5excel%E6%96%87%E4%BB%B6%EF%BC%8C%E6%94%AF%E6%8C%81%E3%80%90.xlsx%20%2F%20.xlsm%20%2F%20.xltx%20%2F%20.xltm%E3%80%91%E6%A0%BC%E5%BC%8F%E7%9A%84%E6%96%87%E4%BB%B6%EF%BC%8C%E5%8F%AF%E5%A4%84%E7%90%86excel%E6%95%B0%E6%8D%AE%E3%80%81%E5%85%AC%E5%BC%8F%E3%80%81%E6%A0%B7%E5%BC%8F%EF%BC%8C%E4%B8%94%E5%8F%AF%E4%BB%A5%E5%9C%A8%E8%A1%A8%E6%A0%BC%E5%86%85%E6%8F%92%E5%85%A5%E5%9B%BE%E8%A1%A8
    
'''
import os
import sys
from typing import List
import logging
import openpyxl
from logging.handlers import RotatingFileHandler


def log_config(level=logging.INFO):
	LOG_FORMAT = '[%(asctime)s][%(levelname)s]: %(message)s'
	logging.basicConfig(level=level, format=LOG_FORMAT)
	# 创建RotatingFileHandler对象,满2MB为一个文件，共备份3个文件
	log_file_handler = RotatingFileHandler(
		filename='test.log', maxBytes=2*1024*1024, backupCount=3)
 	# 设置日志打印格式
	formatter = logging.Formatter(LOG_FORMAT)
	log_file_handler.setFormatter(formatter)
	logging.getLogger('').addHandler(log_file_handler)


def gain_data_from_xlsx(xlsx_file: str) -> List[List[List[str]]]:
	"""Gain data from xlsx

	Args:
		xlsx_file (xslx_file): a file which SheetName == sql_table name, first row is attriable

	Returns:
		List[List[str]]: 
			sheet_attr, sheet_data
				cell_data in each row
	"""
	wb = openpyxl.load_workbook(xlsx_file)
	sheet_name = wb.sheetnames
	all_data = []
	for i in range(0, len(sheet_name)):
		sheet = wb[sheet_name[i]]
		dim = sheet.dimensions
		# get col, row of sheet
		col_num, row_num = sheet[dim.split(":")[1]].column, sheet[dim.split(":")[1]].row
		# read first line as Table Attriable
		attr = []
		attr.append(sheet_name[i])
		for col in range(1, col_num + 1):
			attr.append(sheet.cell(row=1, column=col).value)
		sheet_data = []
		for row in range(2, row_num + 1):
			tmp = []
			for col in range(1, col_num + 1):
				tmp.append(sheet.cell(row=row, column=col).value)
			sheet_data.append(tmp)
		all_data.append(attr)
		all_data.append(sheet_data)
		# tmp = "B1:"+ char(dim.split(":")[1] + 1)
		# cells = sheet[tmp]
		# res = []
		# for row in cells:
		# 	a_row = []
		# 	for col in row:
		# 		a_row.append(col.value)
		# 	res.append(a_row)
		# print(res)
	logging.info('data grain successed!')
	return all_data

def create_sql_into_file(file_location: str, data=List[List[List[str]]]):
	# sql1 = 'INSERT %s (' % table_name
	# sql2 = ','.join(i[0] for i in ls)
	# sql3 = ') VALUES ('
	# sql4 = ','.join('%r' % i[1] for i in ls)
	# sql5 = ');'  
	assert(len(data) % 2 == 0)
 
	attr = []
	for i in range(0, len(data)):
		if i % 2 == 0 :
			attr = data[i]
		else:
			sheet_data = data[i]
			table_name = attr[0]
			logging.debug(f"========== {table_name} =========")
			with open(file_location, 'a+') as FILE:
				# FILE.write(f"DELETE FROM {table_name};\n")
				for row in data[i]:
					if row[0] == None :
						break
					logging.debug("data: %s", row)
					sql = f"INSERT INTO {table_name} values ('"
					sql+= "','".join(str(col) for col in row)
					sql+= "');\n"
					FILE.write(sql)
     	
def clear_file(sql_file):
    with open(sql_file, 'w') as FILE:
        FILE.truncate()

def main(logLevel=logging.INFO):
	log_config(logLevel)
	sql_file = "./insertData.sql"
	xlsx_file = "./SQL_Data.xlsx"

	inp = ""
	print("if you want to using a different xlsx file please input it's path as below: ", end="")
	inp = input()
	if inp != "":
		logging.info(f"your input is :{inp}")
		sql_file = inp
		inp = ""
	inp = ""
	print("if you want to using a different sql output file please input it's path as below: ", end="")
	inp = input()
	if inp != "":
		logging.info(f"your input is :{inp}")
		xlsx_file = inp
		inp = ""
	clear_file(sql_file)
	all_data = gain_data_from_xlsx(xlsx_file)
	create_sql_into_file(sql_file, all_data)

if __name__ == '__main__':
    main()
